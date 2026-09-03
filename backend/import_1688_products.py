# -*- coding: utf-8 -*-
"""
1688商品数据导入脚本
支持1688官方采购助手插件导出的Excel/CSV文件
功能: 自动识别字段 → 标准化 → 选品评分 → 定价分析 → 输出报告

使用方法:
  python import_1688_products.py <文件路径>
  python import_1688_products.py data/户外露营灯.xlsx
  python import_1688_products.py data/露营灯.csv

输出:
  - sourcing_import_results.json (标准化商品数据+选品评分)
  - 控制台表格展示TOP推荐
"""
import sys
import os
import csv
import json
import re
from datetime import datetime

sys.path.insert(0, '.')

from app.services.competitor_analysis_service import compare_pricing

# ============================================
# 字段映射（1688采购助手可能导出的各种字段名）
# ============================================
FIELD_MAPPING = {
    # 商品标题
    "subject": ["商品标题", "商品名称", "宝贝标题", "产品名称", "产品标题",
                "货品标题", "offer标题", "标题", "name", "title", "subject", "product_name"],
    # 1688产品ID
    "product_id": ["产品ID", "商品ID", "offer_id", "product_id", "item_id", "货品ID", "id"],
    # 价格
    "price": ["价格", "批发价", "单价", "售价", "价格区间", "最低价格", "最高价",
              "price", "unit_price", "wholesale_price", "报价", "金额"],
    # 起订量
    "min_order_qty": ["起订量", "最小起订", "起批量", "MOQ", "最小起订量", "起订",
                      "min_order", "min_order_quantity", "起批数量"],
    # 供应商
    "supplier": ["供应商", "公司名称", "店铺名称", "厂家", "供应商名称", "企业名称",
                 "店铺", "company", "supplier", "company_name", "seller", "商家", "工厂"],
    # 商品链接
    "detail_url": ["商品链接", "详情链接", "链接", "URL", "url", "detail_url",
                   "商品URL", "详情页链接", "offer_url", "product_url", "产品链接"],
    # 主图
    "image_url": ["主图", "图片", "商品图", "首图", "image", "image_url", "pic",
                  "主图URL", "图片链接", "product_image", "main_image"],
    # 销量
    "sales": ["销量", "成交", "30天销量", "已售", "销售数量", "月销", "sale_count",
              "sales", "sold", "transaction_count", "成交量", "累计销量"],
    # 复购率
    "repurchase_rate": ["复购率", "repurchase_rate", "回头率", "回购率"],
    # 发货地
    "location": ["发货地", "产地", "工厂地址", "发货地址", "location", "address",
                 "ship_from", "origin", "地区", "省份", "城市", "地址"],
    # SKU/规格
    "sku": ["SKU", "规格", "型号", "sku", "spec", "model", "产品规格", "货号"],
    # 库存
    "stock": ["库存", "库存数量", "stock", "stock_quantity", "可用库存", "现货"],
    # 评分/信用
    "rating": ["评分", "信用", "店铺评分", "rating", "credit_level", "诚信通",
               "店铺等级", "好评率"],
    # 类目
    "category": ["类目", "分类", "category", "类目名称", "产品类目", "一级类目"],
}


def read_file(filepath: str) -> list[dict[str, str]]:
    """读取Excel或CSV文件，返回字典列表"""
    ext = os.path.splitext(filepath)[1].lower()

    if ext == '.csv':
        return read_csv(filepath)
    elif ext in ('.xlsx', '.xls'):
        return read_excel(filepath)
    else:
        raise ValueError(f"不支持的文件格式: {ext}，请使用 .xlsx 或 .csv")


def read_csv(filepath: str) -> list[dict[str, str]]:
    """读取CSV文件，自动检测编码"""
    encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb2312']
    for enc in encodings:
        try:
            with open(filepath, 'r', encoding=enc) as f:
                reader = csv.DictReader(f)
                return [dict(row) for row in reader]
        except (UnicodeDecodeError, UnicodeError):
            continue
    raise ValueError("无法识别CSV文件编码")


def read_excel(filepath: str) -> list[dict[str, str]]:
    """读取Excel文件"""
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    data = []
    for row in rows[1:]:
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        item = {}
        for i, cell in enumerate(row):
            if i < len(headers):
                item[headers[i]] = str(cell).strip() if cell is not None else ''
        data.append(item)

    wb.close()
    return data


def normalize_fields(raw_item: dict[str, str]) -> dict[str, any]:
    """将原始字段映射为标准化字段"""
    normalized = {
        "subject": "",
        "product_id": "",
        "price": None,
        "price_min": None,
        "price_max": None,
        "min_order_qty": None,
        "supplier": "",
        "detail_url": "",
        "image_url": "",
        "sales": None,
        "repurchase_rate": None,
        "location": "",
        "sku": "",
        "stock": None,
        "rating": None,
        "category": "",
        "raw": {},
    }

    # 保存原始数据
    normalized["raw"] = dict(raw_item)

    # 字段映射（先精确匹配，再包含匹配，避免"商品链接"匹配到"商品"）
    for std_field, aliases in FIELD_MAPPING.items():
        # 第一轮：精确匹配
        for raw_key, raw_val in raw_item.items():
            raw_key_clean = str(raw_key).strip().lower()
            for alias in aliases:
                if alias.lower() == raw_key_clean:
                    if raw_val and str(raw_val).strip():
                        normalized[std_field] = str(raw_val).strip()
                    break
        # 第二轮：包含匹配（仅当精确匹配未命中时）
        if not normalized.get(std_field):
            for raw_key, raw_val in raw_item.items():
                raw_key_clean = str(raw_key).strip().lower()
                for alias in aliases:
                    if len(alias) >= 3 and alias.lower() in raw_key_clean:
                        if raw_val and str(raw_val).strip():
                            normalized[std_field] = str(raw_val).strip()
                        break

    # 解析价格（支持区间价 "10.5-25.8" 或 "¥15.80" 或 "-"）
    price_str = str(normalized.get("price", "")).strip()
    if price_str and price_str not in ("-", "—", "暂无", "无", "面议", ""):
        prices = re.findall(r'[\d.]+', price_str)
        if prices:
            price_values = [float(p) for p in prices if float(p) > 0]
            if price_values:
                normalized["price_min"] = min(price_values)
                normalized["price_max"] = max(price_values)
                normalized["price"] = min(price_values)

    # 解析起订量
    moq_str = str(normalized.get("min_order_qty", ""))
    if moq_str:
        nums = re.findall(r'\d+', moq_str)
        if nums:
            normalized["min_order_qty"] = int(nums[0])

    # 解析销量（支持 "已售100+件"、"1.2万"、"12580"、"-" 等格式）
    sales_str = str(normalized.get("sales", "")).strip()
    if sales_str and sales_str not in ("-", "—", "暂无", "无", "0", ""):
        # 处理 "已售100+件" 格式
        if '已售' in sales_str or '件' in sales_str:
            nums = re.findall(r'[\d.]+', sales_str)
            if nums:
                normalized["sales"] = int(float(nums[0]))
            else:
                normalized["sales"] = None
        # 处理 "1.2万" 格式
        elif '万' in sales_str:
            nums = re.findall(r'[\d.]+', sales_str)
            if nums:
                normalized["sales"] = int(float(nums[0]) * 10000)
            else:
                normalized["sales"] = None
        else:
            nums = re.findall(r'\d+', sales_str)
            if nums:
                normalized["sales"] = int(nums[0])
            else:
                normalized["sales"] = None
    else:
        normalized["sales"] = None

    # 解析复购率
    repurchase_str = str(normalized.get("repurchase_rate", ""))
    if repurchase_str:
        nums = re.findall(r'[\d.]+', repurchase_str)
        if nums:
            normalized["repurchase_rate"] = float(nums[0])

    # 产品链接不完整时自动拼接产品ID（如 https://detail.1688.com/offer/ + product_id）
    detail_url = normalized.get("detail_url", "")
    product_id = normalized.get("product_id", "")
    if detail_url and detail_url.endswith("/offer/") and product_id:
        normalized["detail_url"] = detail_url + str(product_id)

    # 解析库存
    stock_str = str(normalized.get("stock", ""))
    if stock_str:
        nums = re.findall(r'\d+', stock_str)
        if nums:
            normalized["stock"] = int(nums[0])

    return normalized


def calculate_sourcing_score(product: dict[str, any]) -> dict[str, any]:
    """
    计算选品评分（0-100分）
    维度: 价格竞争力、销量、起订量友好度、供应商完整度、数据完整度
    """
    scores = {}
    total = 0

    # 1. 价格竞争力 (30分) - 价格越低分越高，但需合理(>1元)
    price = product.get("price")
    if price and price > 1:
        # 户外用品合理价格区间5-100元，越低分越高
        if price <= 10:
            price_score = 30
        elif price <= 30:
            price_score = 25
        elif price <= 50:
            price_score = 18
        elif price <= 100:
            price_score = 12
        else:
            price_score = 6
    else:
        price_score = 5
    scores["价格竞争力"] = price_score
    total += price_score

    # 2. 销量表现 (25分)
    sales = product.get("sales")
    if sales:
        if sales >= 10000:
            sales_score = 25
        elif sales >= 5000:
            sales_score = 22
        elif sales >= 1000:
            sales_score = 18
        elif sales >= 500:
            sales_score = 14
        elif sales >= 100:
            sales_score = 10
        else:
            sales_score = 5
    else:
        sales_score = 3
    scores["销量表现"] = sales_score
    total += sales_score

    # 3. 起订量友好度 (15分) - 起订量越低越适合小批量测试
    moq = product.get("min_order_qty")
    if moq:
        if moq <= 10:
            moq_score = 15
        elif moq <= 50:
            moq_score = 12
        elif moq <= 100:
            moq_score = 9
        elif moq <= 500:
            moq_score = 5
        else:
            moq_score = 2
    else:
        moq_score = 5  # 未知给中等分
    scores["起订量友好度"] = moq_score
    total += moq_score

    # 4. 供应商信息 (15分)
    supplier = product.get("supplier", "")
    location = product.get("location", "")
    rating = product.get("rating")
    supplier_score = 0
    if supplier and len(supplier) > 2:
        supplier_score += 8
    if location:
        supplier_score += 4
    if rating:
        supplier_score += 3
    scores["供应商信息"] = supplier_score
    total += supplier_score

    # 5. 数据完整度 (15分)
    completeness = 0
    fields_to_check = ["subject", "price", "detail_url", "image_url", "supplier"]
    for field in fields_to_check:
        if product.get(field):
            completeness += 3
    scores["数据完整度"] = min(completeness, 15)
    total += min(completeness, 15)

    # 决策建议
    if total >= 80:
        decision = "强烈推荐"
        action = "优先采购样品测试"
    elif total >= 65:
        decision = "推荐"
        action = "可纳入候选池"
    elif total >= 50:
        decision = "一般"
        action = "观望，补充数据后再评估"
    else:
        decision = "不推荐"
        action = "排除"

    return {
        "total_score": total,
        "dimension_scores": scores,
        "decision": decision,
        "action": action,
    }


def analyze_pricing(products: list[dict]) -> dict:
    """对导入商品做定价分析"""
    prices = [p["price"] for p in products if p.get("price") and p["price"] > 0]
    if not prices:
        return {"error": "无有效价格数据"}

    # 用第一个商品的价格作为"我们的价格"参考，分析市场定位
    our_price = prices[0]
    competitor_prices = [{"price": p, "name": f"竞品{i+1}"} for i, p in enumerate(prices[1:10])]

    try:
        analysis = compare_pricing(our_price, competitor_prices)
        return analysis
    except Exception as e:
        return {"error": str(e), "price_count": len(prices),
                "avg_price": round(sum(prices) / len(prices), 2),
                "min_price": min(prices), "max_price": max(prices)}


def main():
    if len(sys.argv) < 2:
        print("用法: python import_1688_products.py <Excel/CSV文件路径>")
        print("示例: python import_1688_products.py data/户外露营灯.xlsx")
        sys.exit(1)

    filepath = sys.argv[1]
    if not os.path.exists(filepath):
        print(f"错误: 文件不存在 - {filepath}")
        sys.exit(1)

    print("=" * 70)
    print("1688商品数据导入 & 选品分析")
    print("=" * 70)

    # Step 1: 读取文件
    print(f"\n--- Step 1: 读取文件 ---")
    print(f"  文件: {filepath}")
    raw_data = read_file(filepath)
    print(f"  原始行数: {len(raw_data)}")
    if raw_data:
        print(f"  原始字段: {list(raw_data[0].keys())[:10]}...")

    if not raw_data:
        print("  错误: 文件为空")
        sys.exit(1)

    # Step 2: 标准化字段
    print(f"\n--- Step 2: 标准化字段 ---")
    products = []
    for i, item in enumerate(raw_data):
        normalized = normalize_fields(item)
        if normalized.get("subject") or normalized.get("detail_url"):
            products.append(normalized)
    print(f"  有效商品数: {len(products)}")
    print(f"  字段识别率: 标题{sum(1 for p in products if p['subject'])}/{len(products)}, "
          f"价格{sum(1 for p in products if p['price'])}/{len(products)}, "
          f"供应商{sum(1 for p in products if p['supplier'])}/{len(products)}")

    # Step 3: 选品评分
    print(f"\n--- Step 3: 选品评分 ---")
    for p in products:
        score_result = calculate_sourcing_score(p)
        p["sourcing_score"] = score_result

    # 按总分排序
    products.sort(key=lambda x: x["sourcing_score"]["total_score"], reverse=True)

    # 打印TOP 10
    print(f"\n  {'排名':<4} {'评分':<6} {'决策':<8} {'价格':<10} {'销量':<8} {'起订':<6} 商品标题")
    print("  " + "-" * 90)
    for i, p in enumerate(products[:10]):
        score = p["sourcing_score"]
        price_str = f"${p['price']:.2f}" if p.get("price") else "N/A"
        sales_str = str(p["sales"]) if p.get("sales") else "N/A"
        moq_str = str(p["min_order_qty"]) if p.get("min_order_qty") else "N/A"
        subject = p["subject"][:35] + "..." if len(p["subject"]) > 35 else p["subject"]
        print(f"  {i+1:<4} {score['total_score']:<6} {score['decision']:<8} "
              f"{price_str:<10} {sales_str:<8} {moq_str:<6} {subject}")

    # Step 4: 定价分析
    print(f"\n--- Step 4: 定价分析 ---")
    pricing = analyze_pricing(products)
    if "error" not in pricing:
        print(f"  市场均价: ${pricing.get('market_avg', 'N/A')}")
        print(f"  价格定位: {pricing.get('price_position', 'N/A')}")
        print(f"  价格百分位: {pricing.get('percentile', 'N/A')}")
        if pricing.get("recommendations"):
            print(f"  建议: {pricing['recommendations'][0] if pricing['recommendations'] else 'N/A'}")
    else:
        prices = [p["price"] for p in products if p.get("price")]
        if prices:
            print(f"  价格区间: ${min(prices):.2f} - ${max(prices):.2f}")
            print(f"  平均价格: ${sum(prices)/len(prices):.2f}")
            print(f"  商品数: {len(prices)}")

    # Step 5: 保存结果
    print(f"\n--- Step 5: 保存结果 ---")
    output = {
        "import_time": datetime.now().isoformat(),
        "source_file": filepath,
        "total_products": len(products),
        "pricing_analysis": pricing,
        "products": products,
        "summary": {
            "avg_score": round(sum(p["sourcing_score"]["total_score"] for p in products) / len(products), 1) if products else 0,
            "strong_recommend": sum(1 for p in products if p["sourcing_score"]["total_score"] >= 80),
            "recommend": sum(1 for p in products if 65 <= p["sourcing_score"]["total_score"] < 80),
            "average": sum(1 for p in products if 50 <= p["sourcing_score"]["total_score"] < 65),
            "not_recommend": sum(1 for p in products if p["sourcing_score"]["total_score"] < 50),
        }
    }

    output_file = "sourcing_import_results.json"
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"  结果已保存: {output_file}")

    # 最终汇总
    print("\n" + "=" * 70)
    print("导入完成汇总")
    print("=" * 70)
    print(f"  导入商品: {len(products)} 个")
    print(f"  平均评分: {output['summary']['avg_score']} 分")
    print(f"  强烈推荐: {output['summary']['strong_recommend']} 个")
    print(f"  推荐: {output['summary']['recommend']} 个")
    print(f"  一般: {output['summary']['average']} 个")
    print(f"  不推荐: {output['summary']['not_recommend']} 个")
    print(f"\n  下一步:")
    print(f"    1. 查看 sourcing_import_results.json 获取完整数据")
    print(f"    2. 对'强烈推荐'商品采购样品测试")
    print(f"    3. 通过 product_listing_service 上架到WooCommerce")
    print(f"    4. 接入 experiment_automation_service 做AB测试")


if __name__ == "__main__":
    main()
